#!/usr/bin/env python3
"""
build.py - Konversi data_input.xlsx ke data.json untuk dashboard Common Enemy UIT JBM
Jalankan: python build.py
"""

import json, sys, re
from datetime import datetime
from pathlib import Path
import openpyxl

EXCEL_FILE = "data_input.xlsx"
JSON_FILE  = "data.json"
HTML_FILE  = "index.html"

UPTS  = ["MALANG","SURABAYA","GRESIK","BALI","MADIUN","PROBOLINGGO"]
CATS  = [
    ("G","G1 AHI Trafo"),("G","G1 Anti Binatang"),("G","G1 Relay Internal"),
    ("G","G2 MV Aparatus"),("G","G3 Switchyard AHI"),("G","G4 GIS"),
    ("G","G5 Common Facility"),
    ("T","T1 Cable Sheath"),("T","T2 SUTT AHI"),("T","T2 ROW B1"),
    ("T","T2 Anti Binatang"),("T","T2 Tapak Tower"),
    ("P","P1 AHI Proteksi"),("P","P1 Design Proteksi"),
    ("P","P2 AHI Catu Daya"),("P","P2 Design Catu Daya"),
]
ROW_KEYS = ["T2 ROW B1"]
ROW_DIV  = 100

# ── Baca tanggal dan data dari Excel ───────────────────────────────────────
wb   = openpyxl.load_workbook(EXCEL_FILE, data_only=True)
ws   = wb["Input Harian"]

# Tanggal di B2
raw_date = ws["B2"].value
if isinstance(raw_date, datetime):
    tanggal = raw_date.strftime("%-d %B %Y")   # e.g. "26 April 2026"
    wl_label = raw_date.strftime("%-d %b")      # e.g. "26 Apr"
elif isinstance(raw_date, str):
    tanggal  = raw_date
    wl_label = raw_date[:6]
else:
    print("ERROR: Tanggal tidak valid di cell B2"); sys.exit(1)

print(f"Tanggal: {tanggal}  |  WL label: {wl_label}")

# Baca data grid: baris 4..19 (16 kategori), kolom per UPT
# Struktur kolom: A=Grup, B=Kategori, C=T_UPT1, D=O_UPT1, E=TL_UPT1, F=T_UPT2, ...
new_data = {}   # {upt: {cat: (t, o, tl)}}
tinjut_today = {}
warnings = []

for ui, upt in enumerate(UPTS):
    new_data[upt] = {}
    tinjut_today[upt] = 0
    for ri, (grp, cat) in enumerate(CATS):
        row  = 4 + ri
        col_t  = 3 + ui * 3   # TARGET column (1-indexed)
        col_o  = col_t + 1
        col_tl = col_t + 2
        t  = int(ws.cell(row=row, column=col_t).value  or 0)
        o  = int(ws.cell(row=row, column=col_o).value  or 0)
        tl = int(ws.cell(row=row, column=col_tl).value or 0)

        # Validasi
        if o > t and t > 0:
            warnings.append(f"⚠  {upt}/{cat}: open({o}) > target({t})")
        if tl > o:
            warnings.append(f"⚠  {upt}/{cat}: tinjut({tl}) > open({o})")

        new_data[upt][cat] = (t, o, tl)
        tinjut_today[upt] += tl

# Tampilkan warning
if warnings:
    print("\nWARNING:")
    for w in warnings: print(" ", w)
    ans = input("\nLanjutkan? (y/n): ").strip().lower()
    if ans != "y":
        print("Dibatalkan."); sys.exit(0)

# ── Baca data.json lama (history) ─────────────────────────────────────────
json_path = Path(JSON_FILE)
if json_path.exists():
    with open(json_path) as f:
        old = json.load(f)
    wl_list = old.get("wl", [])
    old_d   = old.get("d", {})
    print(f"\nHistory: {len(wl_list)} entry (terakhir: {wl_list[-1] if wl_list else '-'})")
else:
    # Pertama kali - tidak ada history
    wl_list = []
    old_d   = {}
    print("\nTidak ada data.json lama — akan dibuat baru.")

# Cek apakah tanggal ini sudah ada
if wl_label in wl_list:
    print(f"\n⚠  Label '{wl_label}' sudah ada di WL.")
    ans = input("Timpa data hari ini? (y/n): ").strip().lower()
    if ans == "y":
        idx = wl_list.index(wl_label)
        # Timpa entry ke-idx
        for upt in UPTS:
            for grp, cat in CATS:
                key = f"{upt}.{grp}.{cat}"
                if key in old_d:
                    old_d[key][idx] = {"t": new_data[upt][cat][0], "o": new_data[upt][cat][1]}
        with open(JSON_FILE, "w") as f:
            json.dump({"wl": wl_list, "last_update": tanggal,
                       "tinjut_today": tinjut_today, "d": old_d}, f)
        print(f"✓ Data {wl_label} ditimpa."); sys.exit(0)
    else:
        print("Dibatalkan."); sys.exit(0)

# Append entry baru
wl_list.append(wl_label)
new_d = old_d.copy()
for upt in UPTS:
    for grp, cat in CATS:
        key = f"{upt}.{grp}.{cat}"
        t, o, tl = new_data[upt][cat]
        if key not in new_d:
            new_d[key] = []
        new_d[key].append({"t": t, "o": o})

# Validasi panjang array
n = len(wl_list)
mismatches = []
for key, arr in new_d.items():
    if len(arr) != n:
        mismatches.append(f"  {key}: {len(arr)} entry (expected {n})")
if mismatches:
    print(f"\n⚠  Array length mismatch ({len(mismatches)} keys):")
    for m in mismatches[:5]: print(m)

# Simpan data.json
out = {
    "wl":           wl_list,
    "last_update":  tanggal,
    "tinjut_today": tinjut_today,
    "d":            new_d
}
with open(JSON_FILE, "w", encoding="utf-8") as f:
    json.dump(out, f, ensure_ascii=False, separators=(",",":"))

print(f"\n✓ data.json diperbarui: {n} entry, last={wl_label}")
print(f"✓ Tinjut hari ini: {tinjut_today}")
print(f"\nLangkah selanjutnya:")
print(f"  git add data.json")
print(f"  git commit -m 'update {wl_label}'")
print(f"  git push")
